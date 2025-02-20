import pandas as pd
from datetime import datetime
from loguru import logger
from pathlib import Path
from core.wallet_manager import Wallet, Chain, TransactionResult
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ResultsTracker:
    def __init__(self, filename: str = "results.xlsx"):
        self.filename = filename
        self.results = []

        # Определяем стили
        self.header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        self.success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Определяем границы
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def update_results(self, wallet: Wallet, chain: Chain, result: TransactionResult):
        # Добавляем результат в память
        self.results.append({
            'Date': datetime.now().strftime('%Y-%m-%d'),
            'Time': datetime.now().strftime('%H:%M:%S'),
            'Wallet Address': wallet.address,
            'Chain': chain.name,
            'Chain ID': chain.id,
            'Module': result.module_name,
            'Status': 'Success' if result.success else 'Failed',
            'Transaction Hash': 'https://blockscout.lisk.com/tx/0x'+result.tx_hash if result.tx_hash else '-',
            'Error': result.error_message if result.error_message else '-'
        })

        try:
            # Создаем DataFrame
            df = pd.DataFrame(self.results)

            # Сохраняем в Excel
            df.to_excel(self.filename, index=False, engine='openpyxl')

            # Применяем форматирование
            self._apply_formatting()

            logger.debug(f"Results updated successfully for address {wallet.address}")

        except Exception as e:
            logger.error(f"Error updating results: {str(e)}")

    def _apply_formatting(self):
        try:
            wb = openpyxl.load_workbook(self.filename)
            ws = wb.active

            # Форматирование заголовков
            for cell in ws[1]:
                cell.fill = self.header_fill
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.border

            # Автоматическая ширина столбцов
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)  # Ограничиваем максимальную ширину
                ws.column_dimensions[column_letter].width = adjusted_width

            # Форматирование данных
            for row in ws.iter_rows(min_row=2):
                # Центрирование всех ячеек
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = self.border

                # Окрашивание строк в зависимости от статуса
                status_cell = row[6]  # Колонка 'Status'
                fill = self.success_fill if status_cell.value == 'Success' else self.error_fill

                for cell in row:
                    cell.fill = fill

                # Форматирование хеша транзакции
                tx_hash_cell = row[7]  # Колонка 'Transaction Hash'
                if tx_hash_cell.value and tx_hash_cell.value != '-':
                    tx_hash_cell.font = Font(color="0000FF", underline="single")

            # Замораживаем верхнюю строку
            ws.freeze_panes = 'A2'

            # Сохраняем изменения
            wb.save(self.filename)

        except Exception as e:
            logger.error(f"Error applying formatting: {str(e)}")

    def get_statistics(self):
        """Получение статистики по результатам"""
        df = pd.DataFrame(self.results)
        if not df.empty:
            stats = {
                'Total Transactions': len(df),
                'Successful': len(df[df['Status'] == 'Success']),
                'Failed': len(df[df['Status'] == 'Failed']),
                'Success Rate': f"{(len(df[df['Status'] == 'Success']) / len(df) * 100):.2f}%",
                'Most Used Chain': df['Chain'].mode().iloc[0] if not df['Chain'].empty else 'N/A',
                'Most Used Module': df['Module'].mode().iloc[0] if not df['Module'].empty else 'N/A',
            }
            return stats
        return None

    def save_results(self):
        """Сохраняет все накопленные результаты и применяет форматирование"""
        try:
            df = pd.DataFrame(self.results)
            df.to_excel(self.filename, index=False)
            self._apply_formatting()

            # Выводим статистику
            stats = self.get_statistics()
            if stats:
                logger.info("Transaction Statistics:")
                for key, value in stats.items():
                    logger.info(f"{key}: {value}")

            logger.info(f"All results saved to {self.filename}")
        except Exception as e:
            logger.error(f"Error saving results: {str(e)}")